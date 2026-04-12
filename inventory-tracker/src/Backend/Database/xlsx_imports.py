import json
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from .models import Item, Recipe, Sales


REQUIRED_HEADERS = {
    "items": {"name", "category", "stock", "cost_price", "price"},
    "sales": {"item", "quantity", "total", "status"},
    "recipes": {"name", "ingredients"},
}


def _normalize_headers(row_values):
    return [str(v).strip().lower() if v is not None else "" for v in row_values]


def _get_sheet_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is required for xlsx imports. Install it with: pip install openpyxl") from exc

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = _normalize_headers(rows[0])
    return headers, rows[1:]


def _to_int(value, default=0):
    if value is None or value == "":
        return default
    return int(value)


def _to_decimal(value, default="0"):
    if value is None or value == "":
        return Decimal(default)
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _row_to_dict(headers, row):
    row_values = list(row)
    if len(row_values) < len(headers):
        row_values += [None] * (len(headers) - len(row_values))
    return {headers[idx]: row_values[idx] for idx in range(len(headers))}


def _is_blank_row(row_dict):
    return all(v in (None, "") for v in row_dict.values())


def _validate_headers(headers, doc_type):
    missing = REQUIRED_HEADERS[doc_type] - set(headers)
    if missing:
        missing_str = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_str}")


def import_items_from_xlsx(uploaded_file):
    headers, rows = _get_sheet_rows(uploaded_file)
    if not headers:
        raise ValueError("The uploaded file is empty.")
    _validate_headers(headers, "items")

    created = 0
    updated = 0
    skipped = 0
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            data = _row_to_dict(headers, row)
            if _is_blank_row(data):
                continue

            try:
                name = str(data.get("name", "")).strip()
                category = str(data.get("category", "")).strip()
                if not name:
                    raise ValueError("name is required")
                if not category:
                    raise ValueError("category is required")

                stock_delta = _to_int(data.get("stock"), default=0)
                cost_price = _to_decimal(data.get("cost_price"), default="0")
                price = _to_decimal(data.get("price"), default="0")

                item = Item.objects.filter(name__iexact=name).first()
                if item:
                    item.category = category or item.category
                    item.stock = item.stock + stock_delta
                    item.cost_price = cost_price
                    item.price = price
                    item.full_clean()
                    item.save()
                    updated += 1
                else:
                    new_item = Item(
                        name=name,
                        category=category,
                        stock=stock_delta,
                        cost_price=cost_price,
                        price=price,
                    )
                    new_item.full_clean()
                    new_item.save()
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": idx, "error": str(exc)})

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def import_sales_from_xlsx(uploaded_file, created_by=None):
    headers, rows = _get_sheet_rows(uploaded_file)
    if not headers:
        raise ValueError("The uploaded file is empty.")
    _validate_headers(headers, "sales")

    created = 0
    skipped = 0
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            data = _row_to_dict(headers, row)
            if _is_blank_row(data):
                continue

            try:
                item_ref = data.get("item")
                quantity = _to_int(data.get("quantity"), default=1)
                if quantity <= 0:
                    raise ValueError("quantity must be greater than 0")

                item = None
                if isinstance(item_ref, (int, float)) or str(item_ref).isdigit():
                    item = Item.objects.filter(id=int(item_ref)).first()
                if item is None and item_ref not in (None, ""):
                    item_name = str(item_ref).strip()
                    item = Item.objects.filter(name__iexact=item_name).first()
                if item is None:
                    raise ValueError("item not found by id or name")

                total_value = data.get("total")
                total = _to_decimal(total_value) if total_value not in (None, "") else (item.price * quantity)

                status = str(data.get("status") or "Completed").strip()
                if status not in {"Pending", "Completed"}:
                    raise ValueError("status must be Pending or Completed")

                sale = Sales.objects.create(
                    item=item,
                    quantity=quantity,
                    total=total,
                    status=status,
                    created_by=created_by,
                )

                date_val = data.get("date")
                if date_val not in (None, ""):
                    if hasattr(date_val, "tzinfo"):
                        sale.date = date_val
                    else:
                        try:
                            parsed = timezone.datetime.fromisoformat(str(date_val))
                            sale.date = parsed
                        except ValueError:
                            pass
                    sale.save(update_fields=["date"])

                created += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": idx, "error": str(exc)})

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }


def _parse_ingredients(raw_ingredients):
    if raw_ingredients in (None, ""):
        raise ValueError("ingredients is required")

    ingredients = []
    text = str(raw_ingredients).strip()

    if text.startswith("["):
        try:
            payload = json.loads(text)
            if not isinstance(payload, list):
                raise ValueError("ingredients JSON must be a list")
            for ing in payload:
                if not isinstance(ing, dict):
                    continue
                item_ref = ing.get("item")
                qty = int(ing.get("quantity", 1))
                item = None
                if isinstance(item_ref, int) or str(item_ref).isdigit():
                    item = Item.objects.filter(id=int(item_ref)).first()
                if item is None and item_ref not in (None, ""):
                    item = Item.objects.filter(name__iexact=str(item_ref).strip()).first()
                if item is None:
                    raise ValueError(f"ingredient item not found: {item_ref}")
                ingredients.append({"item": item.id, "quantity": qty})
            return ingredients
        except json.JSONDecodeError as exc:
            raise ValueError("ingredients JSON is invalid") from exc

    # text format: item:qty;item:qty
    parts = [p.strip() for p in text.split(";") if p.strip()]
    for part in parts:
        item_name, qty_str = (part.split(":", 1) + ["1"])[:2] if ":" in part else (part, "1")
        item_name = item_name.strip()
        qty = int(qty_str.strip())
        item = Item.objects.filter(name__iexact=item_name).first()
        if item is None:
            raise ValueError(f"ingredient item not found: {item_name}")
        ingredients.append({"item": item.id, "quantity": qty})

    if not ingredients:
        raise ValueError("ingredients could not be parsed")

    return ingredients


def import_recipes_from_xlsx(uploaded_file):
    headers, rows = _get_sheet_rows(uploaded_file)
    if not headers:
        raise ValueError("The uploaded file is empty.")
    _validate_headers(headers, "recipes")

    created = 0
    updated = 0
    skipped = 0
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            data = _row_to_dict(headers, row)
            if _is_blank_row(data):
                continue

            try:
                name = str(data.get("name", "")).strip()
                if not name:
                    raise ValueError("name is required")

                ingredients = _parse_ingredients(data.get("ingredients"))

                recipe = Recipe.objects.filter(name__iexact=name).first()
                if recipe:
                    recipe.ingredients = json.dumps(ingredients)
                    recipe.save()
                    updated += 1
                else:
                    Recipe.objects.create(name=name, ingredients=json.dumps(ingredients))
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": idx, "error": str(exc)})

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }