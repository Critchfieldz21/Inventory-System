"""Generate blank xlsx templates for items / sales / recipes imports.

Creates three workbooks under inventory-tracker/templates/ with the exact
column headers the import endpoints validate against, plus a couple of
example rows in each so the expected format is obvious. Re-running this
script overwrites the existing templates.

The data sheet (read by the importer, which uses ``wb.active``) contains
ONLY the header + example rows. Help text lives on a separate
``Instructions`` sheet so it never gets parsed as a data row.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / 'inventory-tracker' / 'templates'

HEADER_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
EXAMPLE_FONT = Font(color='6B7280', italic=True)
INSTRUCTIONS_TITLE_FONT = Font(bold=True, size=14, color='111827')
INSTRUCTIONS_BODY_FONT = Font(size=11, color='374151')


def _build_data_sheet(ws, headers, examples):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for r_offset, row in enumerate(examples, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c_idx, value=value)
            cell.font = EXAMPLE_FONT

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] +
            [len(str(row[col_idx - 1])) for row in examples if col_idx - 1 < len(row)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 14)

    ws.freeze_panes = 'A2'


def _build_instructions_sheet(wb, title, columns, notes):
    ws = wb.create_sheet(title='Instructions')
    ws.cell(row=1, column=1, value=title).font = INSTRUCTIONS_TITLE_FONT
    ws.cell(row=3, column=1, value='Columns').font = Font(bold=True, size=11)
    for i, (name, desc) in enumerate(columns, start=4):
        ws.cell(row=i, column=1, value=name).font = Font(bold=True, size=11)
        ws.cell(row=i, column=2, value=desc).font = INSTRUCTIONS_BODY_FONT

    notes_start = 4 + len(columns) + 2
    ws.cell(row=notes_start, column=1, value='Notes').font = Font(bold=True, size=11)
    for i, note in enumerate(notes, start=notes_start + 1):
        ws.cell(row=i, column=1, value=f'• {note}').font = INSTRUCTIONS_BODY_FONT
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=4)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70


def build_items_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Items'

    headers = ['name', 'category', 'stock', 'cost_price', 'price']
    examples = [
        ['Beef Patty', 'Meat', 50, 1.25, 3.00],
        ['Brioche Bun', 'Bakery', 80, 0.30, 0.90],
        ['Cheddar Slice', 'Dairy', 60, 0.20, 0.75],
    ]
    _build_data_sheet(ws, headers, examples)
    _build_instructions_sheet(
        wb,
        title='Items import',
        columns=[
            ('name',       'Display name. Required. Existing items with the same name (case-insensitive) are updated, not duplicated.'),
            ('category',   'Free-text category, e.g. Meat, Bakery, Dairy. Required.'),
            ('stock',      'Quantity in stock. For existing items this is ADDED to the current stock (treated as a restock).'),
            ('cost_price', 'What you pay for one unit. Must be strictly less than price.'),
            ('price',      'What you charge customers per unit.'),
        ],
        notes=[
            'Replace the example rows with your own data — keep the header row intact.',
            'cost_price MUST be less than price; rows that violate this are skipped.',
            'Empty rows are ignored, so you can leave gaps.',
        ],
    )
    wb.save(path)


def build_sales_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales'

    headers = ['item', 'quantity', 'total', 'status', 'date']
    # Dates are spread across the last 90 days (rolling) from 2026-05-07 so
    # that all three dashboard views (Last 7 days / 30 days / 13 weeks) light
    # up with data after import. Cumulative units stay well under the stock
    # provided by items_template.xlsx (50/80/60 for patty/bun/cheese).
    examples = [
        # ── Last 7 days (May 1 — May 7) ─────────────────────────────────
        ['Beef Patty',    2, 6.00,  'Completed', '2026-05-07T11:30:00'],
        ['Brioche Bun',   3, 2.70,  'Completed', '2026-05-07T13:15:00'],
        ['Beef Patty',    3, 9.00,  'Completed', '2026-05-06T12:45:00'],
        ['Cheddar Slice', 4, 3.00,  'Completed', '2026-05-05T18:00:00'],
        ['Beef Patty',    2, 6.00,  'Completed', '2026-05-04T19:30:00'],
        ['Brioche Bun',   4, 3.60,  'Completed', '2026-05-01T14:10:00'],
        # ── Older than 7d, within last 30d (Apr 8 — Apr 30) ─────────────
        ['Beef Patty',    4, 12.00, 'Completed', '2026-04-28T13:00:00'],
        ['Cheddar Slice', 3, 2.25,  'Completed', '2026-04-22T17:20:00'],
        ['Brioche Bun',   5, 4.50,  'Completed', '2026-04-15T12:00:00'],
        # ── Older than 30d, within last 90d (Feb 7 — Apr 7) ─────────────
        ['Beef Patty',    3, 9.00,  'Completed', '2026-04-05T11:45:00'],
        ['Brioche Bun',   6, 5.40,  'Completed', '2026-03-22T13:30:00'],
        ['Beef Patty',    4, 12.00, 'Completed', '2026-03-08T14:00:00'],
        ['Cheddar Slice', 2, 1.50,  'Completed', '2026-02-18T12:30:00'],
        # ── Pending (no date — uses time of import) ─────────────────────
        ['Cheddar Slice', 1, 0.75,  'Pending',   ''],
    ]
    _build_data_sheet(ws, headers, examples)
    _build_instructions_sheet(
        wb,
        title='Sales import',
        columns=[
            ('item',     'Either the item NAME (case-insensitive) or its numeric ID. The item must already exist in your inventory.'),
            ('quantity', 'Units sold. Must be a positive integer. Stock is automatically deducted on import.'),
            ('total',    'Total sale amount. Optional — leave blank to auto-calculate as item.price × quantity.'),
            ('status',   'Either "Completed" or "Pending". Defaults to "Completed".'),
            ('date',     'Optional ISO-8601 timestamp, e.g. 2026-05-01T12:30:00. Defaults to the time of import.'),
        ],
        notes=[
            'Item must exist in YOUR inventory before importing the sale.',
            'Imported sales decrement inventory stock just like sales added through the UI.',
            'Example dates here are spread across the last 90 days (today = 2026-05-07) so all three dashboard views populate.',
        ],
    )
    wb.save(path)


def build_recipes_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recipes'

    headers = ['name', 'ingredients']
    examples = [
        ['Cheeseburger', 'Beef Patty:1; Brioche Bun:1; Cheddar Slice:1'],
        ['Double Cheeseburger', 'Beef Patty:2; Brioche Bun:1; Cheddar Slice:2'],
    ]
    _build_data_sheet(ws, headers, examples)
    _build_instructions_sheet(
        wb,
        title='Recipes import',
        columns=[
            ('name',        'Recipe name. Existing recipes with the same name are updated, not duplicated.'),
            ('ingredients', 'Semi-colon-separated list of "ItemName:qty" pairs, e.g. "Beef Patty:1; Brioche Bun:1".'),
        ],
        notes=[
            'Every ingredient item must already exist in your inventory — import items first.',
            'JSON is also accepted: [{"item":"Beef Patty","quantity":1}, {"item":"Brioche Bun","quantity":1}].',
            'Quantities are integers; "Beef Patty:2" means the recipe uses 2 patties.',
        ],
    )
    wb.save(path)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    targets = [
        (OUT_DIR / 'items_template.xlsx', build_items_template),
        (OUT_DIR / 'sales_template.xlsx', build_sales_template),
        (OUT_DIR / 'recipes_template.xlsx', build_recipes_template),
    ]
    for path, builder in targets:
        builder(path)
        print(f'wrote {path}')


if __name__ == '__main__':
    main()
